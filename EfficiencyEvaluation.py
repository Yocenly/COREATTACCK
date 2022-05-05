import matplotlib.pyplot as plt
import numpy as np

from __header__ import *
from openpyxl import load_workbook
img_cache = "./cache/img/"


if __name__ == "__main__":
    filenames = ["./cache/excel/coreAttack_GreedyCOREATTACK.xlsx",
                 "./cache/excel/coreAttack_COREATTACK.xlsx",
                 "./cache/excel/coreAttack_RED.xlsx"]
    plt.rc('font', family='Times New Roman')
    plt.figure(figsize=(18, 6))
    plt.rcParams['xtick.direction'] = 'in'
    plt.rcParams['ytick.direction'] = 'in'
    ax = plt.axes()
    ax.set_ylim(10 ** -2, 10 ** 5)
    ax.set_yscale('log')
    # ax.spines['left'].set_linewidth(2)
    # ax.spines['bottom'].set_linewidth(2)
    # ax.spines['top'].set_visible(False)
    # ax.spines['right'].set_visible(False)
    wb = load_workbook(filenames[0])
    sheet = wb["Sheet1"]
    names = []
    time_greedy = []
    for i in range(2, sheet.max_row):
        names.append(sheet.cell(row=i, column=1).value[2:])
    for i in range(2, sheet.max_row):
        time_greedy.append(sheet.cell(row=i, column=sheet.max_column).value)

    wb = load_workbook(filenames[1])
    sheet = wb["Sheet1"]
    time_none = []
    for i in range(2, sheet.max_row):
        time_none.append(sheet.cell(row=i, column=sheet.max_column).value)

    wb = load_workbook(filenames[2])
    sheet = wb["Sheet1"]
    time_random = []
    for i in range(2, sheet.max_row):
        time_random.append(sheet.cell(row=i, column=sheet.max_column).value)

    x = np.arange(len(names))
    line = np.zeros(len(names) + 1)
    xline = np.arange(len(names))
    xline = np.append(xline, 15)
    print(xline)
    width = 0.25

    plt.bar(x - width, height=time_random, width=width, alpha=0.9, label="RED")
    plt.bar(x, height=time_none, width=width, alpha=0.9, label="COREATTACK")
    plt.bar(x + width, height=time_greedy, width=width, alpha=0.9, label="GreedyCOREATTACK")
    # plt.ylim(0, 1.2)
    plt.plot(xline-0.5, line, color="black", linewidth=1)
    plt.ylabel("Time Consumption (seconds)", size=24)

    # ax.yaxis.set_major_locator(ticker.LogLocator(base=100.0, numticks=5))

    plt.xticks(x, names, rotation=-15)
    # plt.xlabel("Netwroks", size=24)
    plt.legend(fontsize=24)  # 设置题注
    plt.tick_params(labelsize=20)

    # ax.spines['bottom'].set_visible(False)
    # ax.spines['left'].set_visible(False)
    plt.tight_layout()
    plt.savefig(img_cache + "time_consumption.pdf", format='pdf')
    plt.show()



    # # 低时间消耗的图
    # plt.rc('font', family='Times New Roman')
    # plt.figure(1)
    # names = ["Cora", "OpenFlights", "Autonomous", "Mangwet", "Florida"]
    # time_none = [0.3302, 0.1785, 0.2723, 0.0163, 0.0242]
    # time_greedy = [0.3397, 0.3623, 0.3623, 0.0401, 0.0383]
    # time_random = [0.409, 1.1167, 0.6276, 0.1405, 0.1524]
    #
    # x = np.arange(len(names))
    # width = 0.2
    #
    # plt.bar(x - width, height=time_random, width=width, alpha=0.8, label="RED")
    # plt.bar(x, height=time_none, width=width, alpha=0.8, label="COREATTACK")
    # plt.bar(x + width, height=time_greedy, width=width, alpha=0.8, label="Greedy")
    # plt.ylim(0, 1.2)
    # plt.ylabel("Time Consumption (seconds)", size=20)
    #
    # plt.xticks(x, names, rotation=-10)
    # plt.xlabel("Netwrok", size=20)
    # plt.legend(fontsize=18)  # 设置题注
    # plt.tick_params(labelsize=16)
    #
    # plt.tight_layout()
    # plt.savefig(img_cache + "low_time.pdf", format='pdf')
    # plt.show()

    # # 高时间消耗的图
    # plt.figure(2)
    # names = ["Brightkite", "Patents", "Gowalla", "Enron-Email", "Web"]
    # time_none = [0.9746, 102.4644, 4.9304, 1.3946, 4.9855]
    # time_greedy = [1.2531, 101.81, 4.7838, 4.9311, 4.8642]
    # time_random = [4.4975, 108.7764, 6.3535, 19.8735, 4.9975]
    #
    # x = np.arange(len(names))
    # width = 0.2
    #
    # plt.bar(x - width, height=time_random, width=width, alpha=0.8, label="RED")
    # plt.bar(x, height=time_none, width=width, alpha=0.8, label="COREATTACK")
    # plt.bar(x + width, height=time_greedy, width=width, alpha=0.8, label="Greedy")
    #
    # plt.ylim(0, 120)
    # plt.ylabel("Time Consumption (seconds)", size=20)
    #
    # plt.xticks(x, names, rotation=-10)
    # plt.xlabel("Netwrok", size=20)
    # plt.legend(fontsize=18)  # 设置题注
    # plt.tick_params(labelsize=16)
    #
    # plt.tight_layout()
    # plt.savefig(img_cache + "high_time.pdf", format='pdf')
    # plt.show()
    # pass



